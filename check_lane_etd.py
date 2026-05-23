import openpyxl
from datetime import datetime

wb = openpyxl.load_workbook(r'C:\Users\elaineteh\WorkBuddy\20260521093934\data\daily booking.xlsx', data_only=True)
print("Sheets:", wb.sheetnames)
ws = wb[wb.sheetnames[0]]

def s(v):
    if v is None: return ''
    if isinstance(v, str): return v.strip()
    return str(v).strip()

def dt(val):
    if val is None: return ''
    if isinstance(val, datetime): return val.strftime('%Y-%m-%d')
    return s(val)

lane_etds = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    trunk_vessel = s(row[3])
    if not trunk_vessel or trunk_vessel == '0000E': continue
    lane = s(row[8])      # col I: TRUNK LANE
    etd = dt(row[16])     # col Q: ETD
    if not lane: continue
    if lane not in lane_etds:
        lane_etds[lane] = []
    if etd:
        lane_etds[lane].append(etd)

# Show stats
for lane in sorted(lane_etds.keys()):
    etds = sorted(set(lane_etds[lane]))
    print(f"\nLane: {lane}")
    print(f"  Unique ETDs: {len(etds)}")
    print(f"  Range: {etds[0]} ~ {etds[-1]}")
    print(f"  All: {etds[:10]}{'...' if len(etds) > 10 else ''}")
